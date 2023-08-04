import json
import openpyxl


def convert_excel_to_json(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for index, cell_value in enumerate(row):
            header = sheet.cell(row=1, column=index + 1).value
            row_data[header] = cell_value
        data.append(row_data)
    print(data)
    json_data = json.dumps(data, indent=4)
    return json_data
